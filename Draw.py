import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import matplotlib.colors as mcolors
from openpyxl import load_workbook
from openpyxl import Workbook
book=Workbook()
data=book.active
dfrow = load_workbook(filename="TirResult10.xlsx",  data_only=True)
df = dfrow.active
def col(c,df):
    col=[]
    for i in range(2,16):
        col.append(df.cell(row=i,column=c).value)
    return col
n=col(1,df)
setup=col(2,df)
ppsize=col(3,df)
keygentime=col(4,df)
pksize=col(5,df)
keyupdatetime=col(16,df)
keyupdatesize=col(7,df)
keyverify=col(8,df)
batch=col(9,df)
encryptiontime=col(10,df)
ciphertext=col(11,df)
encryptiontimeROM=col(12,df)
ciphertextROM=col(13,df)
decryptiontime=col(14,df)
decryptiontimeROM=col(15,df)

fig, ((ax0, ax1, ax2),(ax3, ax4, ax5)) = plt.subplots(nrows=2, ncols=3, sharex=True,
                                    figsize=(12, 6))
ax0.set_title('Key Verification time')
ax0.errorbar(n,keyverify,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=8, label='Without Batching')
ax0.grid()
ax0.errorbar(n,batch,color='darkgreen', linestyle='dashed', linewidth = 2,
         marker='s', markerfacecolor='c', markersize=8, label='With Batching')
ax0.legend(loc=2,prop={'size': 12})
ax0.set_xlabel('Number of updates')
ax0.set_ylabel('Time (sec)')

ax1.set_title('Key updating time')
ax1.plot(n,keyupdatetime,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=8)
ax1.grid()

ax1.set_xlabel('Number of updates')
ax1.set_ylabel('Time (ms)')



ax2.set_title('Transcript size')
ax2.plot(n,keyupdatesize,color='purple', linestyle='dashed', linewidth = 2,
         marker='+', markerfacecolor='gold', markersize=8)
ax2.grid()

ax2.set_xlabel('Number of updates')
ax2.set_ylabel('Size (kbyte)')


ax3.set_title('Encryption time')
ax3.errorbar(n,encryptiontime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=8, label='Pairing-based')
ax3.grid()
ax3.errorbar(n,encryptiontimeROM,color='purple', linestyle='dashed', linewidth = 2,
         marker='d', markerfacecolor='b', markersize=8, label='Hash-based')
ax3.legend(loc=2,prop={'size': 12})
ax3.set_xlabel('Number of updates')
ax3.set_ylabel('Time (ms)')
ax3.set_ylim([0,45])

ax4.set_title('Decryption time')
ax4.errorbar(n,decryptiontime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=8, label='Pairing-based')
ax4.grid()
ax4.errorbar(n,decryptiontimeROM,color='purple', linestyle='dashed', linewidth = 2,
         marker='d', markerfacecolor='b', markersize=8, label='Hash-based')
ax4.legend(loc=2,prop={'size': 12})
ax4.set_xlabel('Number of updates')
ax4.set_ylabel('Time (ms)')
ax4.set_ylim([0,7])

ax5.set_title('Ciphertext size')
ax5.errorbar(n,ciphertext,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=8, label='Pairing-based')
ax5.grid()
ax5.errorbar(n,ciphertextROM,color='purple', linestyle='dashed', linewidth = 2,
         marker='d', markerfacecolor='b', markersize=8, label='Hash-based')
ax5.legend(loc=2,prop={'size': 12})
ax5.set_xlabel('Number of updates')
ax5.set_ylabel('Size (byte)')
ax5.set_ylim([0,1500])
plt.draw()
plt.savefig("Perf.pdf", bbox_inches='tight')